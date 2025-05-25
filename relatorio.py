# -*- coding: utf-8 -*-
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import datetime
import os # Importar os para verificar a fonte

class PDFRelatorio(FPDF):
    def __init__(self, orientation="P", unit="mm", format="A4", titulo="Relatório"):
        super().__init__(orientation, unit, format)
        self.titulo_relatorio = titulo
        # Usar fonte padrão Arial para maior portabilidade
        self.set_font("Arial", size=10)
        # Assumir que a fonte padrão pode não suportar UTF-8 completamente sem configuração adicional
        self.supports_utf8 = False

    def _encode_str(self, text):
        """Codifica a string para Latin-1 se UTF-8 não for suportado."""
        if self.supports_utf8:
            return str(text) # FPDF com uni=True lida com UTF-8
        else:
            return str(text).encode("latin-1", "replace").decode("latin-1")

    def header(self):
        if self.supports_utf8:
            self.set_font(self.font_name, "B", 12)
        else:
            self.set_font("Arial", "B", 12)
        self.cell(0, 10, self._encode_str(self.titulo_relatorio), 0, 1, "C")

        if self.supports_utf8:
            self.set_font(self.font_name, "I", 8)
        else:
            self.set_font("Arial", "I", 8)
        gerado_em = f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        self.cell(0, 5, self._encode_str(gerado_em), 0, 1, "C")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        if self.supports_utf8:
            self.set_font(self.font_name, "I", 8)
        else:
            self.set_font("Arial", "I", 8)
        pagina_str = f"Página {self.page_no()}/{{nb}}"
        self.cell(0, 10, self._encode_str(pagina_str), 0, 0, "C")

    def chapter_title(self, title):
        if self.supports_utf8:
            self.set_font(self.font_name, "B", 11)
        else:
            self.set_font("Arial", "B", 11)
        self.cell(0, 6, self._encode_str(title), 0, 1, "L")
        self.ln(4)

    def chapter_body(self, colunas, dados):
        if not dados or not colunas:
            if self.supports_utf8:
                self.set_font(self.font_name, "I", 10)
            else:
                self.set_font("Arial", "I", 10)
            self.cell(0, 10, self._encode_str("Nenhum dado encontrado para este relatório."), 0, 1)
            return

        # Calcular larguras das colunas (distribuição simples)
        num_cols = len(colunas)
        largura_util = self.w - 2 * self.l_margin
        largura_col = largura_util / num_cols

        # Cabeçalho da Tabela
        if self.supports_utf8:
            self.set_font(self.font_name, "B", 9)
        else:
            self.set_font("Arial", "B", 9)
        self.set_fill_color(200, 220, 255) # Azul claro
        for i, col in enumerate(colunas):
            self.cell(largura_col, 7, self._encode_str(str(col)), 1, 0, "C", 1)
        self.ln()

        # Dados da Tabela
        if self.supports_utf8:
            self.set_font(self.font_name, "", 8)
        else:
            self.set_font("Arial", "", 8)
        self.set_fill_color(255, 255, 255)
        fill = False
        for linha in dados:
            # Usar as colunas fornecidas para garantir a ordem e extrair valores
            for i, col_key in enumerate(colunas):
                valor = linha.get(col_key) # Tenta pegar pela chave exata
                # Se a chave não existir (ex: 'Total Vendido' vs 'total_vendido'), tentar case-insensitive ou por índice?
                # Por simplicidade, vamos assumir que as chaves em 'dados' correspondem a 'colunas'
                if valor is None:
                     # Tentar encontrar chave ignorando case/underline?
                     found = False
                     for k, v in linha.items():
                         if str(k).strip().lower().replace("_","") == str(col_key).strip().lower().replace("_",""):
                             valor = v
                             found = True
                             break
                     if not found:
                         valor = "N/A" # Valor padrão se não encontrar

                # Formatar valores específicos
                if isinstance(valor, float):
                    texto_celula = f"R$ {valor:.2f}"
                elif isinstance(valor, datetime.datetime):
                    texto_celula = valor.strftime("%d/%m/%Y %H:%M:%S")
                elif isinstance(valor, datetime.date):
                    texto_celula = valor.strftime("%d/%m/%Y")
                else:
                    texto_celula = str(valor if valor is not None else "")

                # Truncar texto se for muito longo
                texto_encoded_para_largura = self._encode_str(texto_celula)
                if self.get_string_width(texto_encoded_para_largura) > largura_col - 2:
                    while self.get_string_width(self._encode_str(texto_celula + "...")) > largura_col - 2 and len(texto_celula) > 0:
                        texto_celula = texto_celula[:-1]
                    texto_celula += "..."

                texto_final_encoded = self._encode_str(texto_celula)
                self.cell(largura_col, 6, texto_final_encoded, 1, 0, "L", fill)
            self.ln()
            fill = not fill

def exportar_para_pdf(titulo, colunas, dados, nome_arquivo):
    """Exporta os dados de um relatório para um arquivo PDF."""
    try:
        pdf = PDFRelatorio(titulo=titulo)
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.chapter_body(colunas, dados)
        pdf.output(nome_arquivo, "F")
        return True
    except Exception as e:
        print(f"Erro ao gerar PDF: {e}")
        # Tentar gerar sem fonte UTF-8 como fallback?
        try:
            pdf = PDFRelatorio(titulo=titulo)
            pdf.supports_utf8 = False # Forçar fallback para Arial/Latin-1
            pdf.set_font("Arial", size=10)
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.chapter_body(colunas, dados)
            pdf.output(nome_arquivo, "F")
            print("PDF gerado com fonte Arial (fallback).")
            return True
        except Exception as e2:
            print(f"Erro ao gerar PDF (fallback Arial): {e2}")
            return False

def exportar_para_excel(titulo, colunas, dados, nome_arquivo):
    """Exporta os dados de um relatório para um arquivo Excel (.xlsx)."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:31] # Limita o nome da aba

        # Título do Relatório (mesclado)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
        cell_titulo = ws.cell(row=1, column=1, value=titulo)
        cell_titulo.font = Font(bold=True, size=14)
        cell_titulo.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 20

        # Cabeçalhos das Colunas
        for col_idx, header in enumerate(colunas, 1):
            cell_header = ws.cell(row=3, column=col_idx, value=header)
            cell_header.font = Font(bold=True)
            cell_header.alignment = Alignment(horizontal="center")

        # Dados
        row_num = 4
        for linha_dict in dados:
            col_num = 1
            # Garantir que os dados sejam escritos na ordem das colunas
            for col_key in colunas:
                 valor = None
                 # Tentar encontrar a chave exata ou com variações comuns
                 possible_keys = [col_key, str(col_key).lower(), str(col_key).replace(" ", "_").lower()]
                 for pk in possible_keys:
                     if pk in linha_dict:
                         valor = linha_dict[pk]
                         break
                 # Se ainda não achou, tentar pelo nome da coluna no dict (case insensitive)
                 if valor is None:
                     for key_dict, val_dict in linha_dict.items():
                         if str(key_dict).strip().lower() == str(col_key).strip().lower():
                             valor = val_dict
                             break

                 cell = ws.cell(row=row_num, column=col_num, value=valor)
                 # Formatação básica
                 if isinstance(valor, (int, float)):
                     cell.number_format = "#,##0.00" if isinstance(valor, float) else "0"
                     cell.alignment = Alignment(horizontal="right")
                 elif isinstance(valor, (datetime.datetime, datetime.date)):
                     cell.number_format = "dd/mm/yyyy hh:mm" if isinstance(valor, datetime.datetime) else "dd/mm/yyyy"
                     cell.alignment = Alignment(horizontal="center")
                 else:
                     cell.alignment = Alignment(horizontal="left")
                 col_num += 1
            row_num += 1

        # Ajustar largura das colunas
        for col_idx, _ in enumerate(colunas, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            column = ws[column_letter]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(nome_arquivo)
        return True
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        return False

# Exemplo de uso (para teste)
if __name__ == "__main__":
    colunas_teste = ["ID", "Nome", "Valor", "Data"]
    dados_teste = [
        {"ID": 1, "Nome": "Item A", "Valor": 123.45, "Data": datetime.date(2024, 5, 20)},
        {"ID": 2, "Nome": "Item B com nome longo", "Valor": 987.60, "Data": datetime.date(2024, 5, 21)},
        {"ID": 3, "Nome": "Item C", "Valor": 50.00, "Data": datetime.date(2024, 5, 22)}
    ]
    titulo_teste = "Relatório de Teste (Maio 2024)"

    # Definir caminhos de saída na pasta do projeto
    output_dir = "/home/ubuntu/sistema_loja"
    pdf_path = os.path.join(output_dir, "relatorio_teste.pdf")
    excel_path = os.path.join(output_dir, "relatorio_teste.xlsx")

    print("Gerando PDF de teste...")
    if exportar_para_pdf(titulo_teste, colunas_teste, dados_teste, pdf_path):
        print(f"PDF gerado com sucesso: {pdf_path}")
    else:
        print("Falha ao gerar PDF.")

    print("\nGerando Excel de teste...")
    if exportar_para_excel(titulo_teste, colunas_teste, dados_teste, excel_path):
        print(f"Excel gerado com sucesso: {excel_path}")
    else:
        print("Falha ao gerar Excel.")

